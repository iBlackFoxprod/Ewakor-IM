from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Product, StockMovement, IncomingStock, Settings
from datetime import datetime
import os
from flask_babel import Babel
from werkzeug.utils import secure_filename
from PIL import Image
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import UserMixin
from flask_babel import gettext as _, lazy_gettext as _l
import csv
from io import TextIOWrapper, StringIO, BytesIO
import requests
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired
import shutil
from functools import wraps
import threading
import time
from sqlalchemy import or_
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Get the absolute path to the instance directory
basedir = os.path.abspath(os.path.dirname(__file__))
instance_path = os.path.join(basedir, 'instance')
if not os.path.exists(instance_path):
    os.makedirs(instance_path)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(instance_path, "ewakor.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['BABEL_DEFAULT_LOCALE'] = 'fr'
app.config['BABEL_TRANSLATION_DIRECTORIES'] = 'translations'
app.config['LANGUAGES'] = ['en', 'fr']

# Ensure upload directories exist
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'logos'), exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'products'), exist_ok=True)

# Initialize Babel
babel = Babel(app)

BACKUP_FOLDER = os.path.join(basedir, 'backups')
if not os.path.exists(BACKUP_FOLDER):
    os.makedirs(BACKUP_FOLDER)

# Initialize database
db.init_app(app)
with app.app_context():
    db.create_all()

# Admin-only decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash(_('Admin access required'), 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@babel.localeselector
def get_locale():
    return 'fr'

login_manager = LoginManager()
login_manager.init_app(app)
setattr(login_manager, 'login_view', 'login')

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        email = form.username.data
        password = form.password.data
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash(_('Invalid email or password'), 'danger')
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

def product_to_dict(product):
    return {
        'id': product.id,
        'name': product.name,
        'category': product.category,
        'product_type': product.product_type,
        'container_type': product.container_type,
        'quantity': product.quantity,
        'unit': product.unit,
        'image_path': product.image_path
    }

# Dashboard
@app.route('/')
@login_required
def dashboard():
    products = Product.query.all()
    low_stock = Product.query.filter(Product.quantity <= Product.min_stock_level).all()
    incoming = IncomingStock.query.filter_by(status='pending').all()
    settings = Settings.query.first()
    products_json = [product_to_dict(p) for p in products]
    return render_template(
        'dashboard.html',
        products=products,
        products_json=products_json,
        low_stock=low_stock,
        incoming=incoming,
        settings=settings
    )

# Settings
@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if current_user.role != 'admin':
        flash(_('Access denied. Admin privileges required.'), 'danger')
        return redirect(url_for('dashboard'))
    
    settings = Settings.query.first()
    if not settings:
        settings = Settings()
        db.session.add(settings)
        db.session.commit()
    
    if request.method == 'POST':
        if 'logo' in request.files:
            logo_path = save_image(request.files['logo'], 'logos')
            if logo_path:
                print(f"DEBUG: Logo path generated by save_image: {logo_path}")
                settings.logo_path = logo_path
        
        settings.language = request.form.get('language', 'en')
        db.session.commit()
        flash(_('Settings updated successfully.'), 'success')
        return redirect(url_for('settings'))
    
    return render_template('settings.html', settings=settings)

# Product management
@app.route('/products')
@login_required
def products():
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    query = Product.query
    if search:
        query = query.filter(Product.name.ilike(f'%{search}%'))
        products = query.all()
        pagination = None
    else:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        products = pagination.items
    settings = Settings.query.first()
    products_list = [product_to_dict(p) for p in products]
    return render_template('products.html', products=products, products_json=products_list, settings=settings, pagination=pagination, search=search)

@app.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if request.method == 'POST':
        name = request.form.get('name')
        category = request.form.get('category')
        product_type = request.form.get('product_type')
        container_type = request.form.get('container_type')
        quantity = int(request.form.get('quantity', 0))
        min_stock_level = int(request.form.get('min_stock_level', 10))
        unit = request.form.get('unit', 'pcs')
        
        image_path = None
        if 'image' in request.files:
            image = request.files['image']
            if image and image.filename:
                image_path = save_image(image, 'products')
        
        product = Product(
            name=name,
            category=category,
            product_type=product_type,
            container_type=container_type,
            quantity=quantity,
            min_stock_level=min_stock_level,
            unit=unit,
            image_path=image_path
        )
        
        db.session.add(product)
        db.session.commit()
        
        backup_db()
        
        flash(_('Product added successfully'), 'success')
        return redirect(url_for('products'))
    
    return render_template('add_product.html')

@app.route('/products/edit/<int:product_id>', methods=['GET', 'POST'])
@login_required
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        product.name = request.form.get('name')
        product.category = request.form.get('category')
        product.product_type = request.form.get('product_type')
        product.container_type = request.form.get('container_type')
        product.quantity = int(request.form.get('quantity', 0))
        product.min_stock_level = int(request.form.get('min_stock_level', 10))
        product.unit = request.form.get('unit', 'pcs')
        
        if 'image' in request.files:
            image = request.files['image']
            if image and image.filename:
                # Delete old image if exists
                if product.image_path and isinstance(product.image_path, str):
                    try:
                        os.remove(os.path.join(str(app.static_folder), str(product.image_path)))
                    except Exception:
                        pass
                product.image_path = save_image(image, 'products')
        
        db.session.commit()
        backup_db()
        flash(_('Product updated successfully'), 'success')
        return redirect(url_for('products'))
    
    return render_template('edit_product.html', product=product)

@app.route('/products/delete/<int:product_id>', methods=['POST'])
@login_required
def delete_product(product_id):
    product = Product.query.get_or_404(product_id)
    
    # Delete product image if exists
    if product.image_path and isinstance(product.image_path, str):
        try:
            os.remove(os.path.join(str(app.static_folder), str(product.image_path)))
        except Exception:
            pass
    
    db.session.delete(product)
    db.session.commit()
    
    backup_db()
    
    flash(_('Product deleted successfully'), 'success')
    return redirect(url_for('products'))

@app.route('/products/upload-csv', methods=['POST'])
@login_required
def upload_products_csv():
    if 'csv_file' not in request.files and 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('products'))
    file = request.files.get('csv_file') or request.files.get('file')
    if not file or file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('products'))
    try:
        stream = TextIOWrapper(file.stream, encoding='utf-8')
        reader = csv.DictReader(stream)
        count = 0
        for row in reader:
            name = row.get('Product Name') or row.get('name')
            quantity = int(row.get('Quantity', 0) or row.get('quantity', 0))
            if not name:
                continue
            product = Product(
                name=name,
                quantity=quantity,
                product_type='other',
                container_type='',
                min_stock_level=1,
                unit='pcs'
            )
            db.session.add(product)
            count += 1
        db.session.commit()
        backup_db()
        flash(f'Successfully uploaded {count} products from CSV.', 'success')
    except Exception as e:
        flash(f'Error uploading CSV: {e}', 'danger')
    return redirect(url_for('products'))

# Stock movements
@app.route('/stock/move', methods=['POST'])
@login_required
def move_stock():
    product_id = request.form['product_id']
    movement_type = request.form['movement_type']
    quantity = int(request.form['quantity'])
    reference = request.form['reference']
    notes = request.form['notes']
    
    product = Product.query.get_or_404(product_id)
    if movement_type == 'out' and product.quantity < quantity:
        return jsonify({'error': _('Insufficient stock')}), 400
    
    movement = StockMovement(
        product_id=product_id,
        movement_type=movement_type,
        quantity=quantity,
        reference=reference,
        notes=notes,
        created_by=current_user.id
    )
    
    if movement_type == 'in':
        product.quantity += quantity
    else:
        product.quantity -= quantity
    
    db.session.add(movement)
    db.session.commit()
    return jsonify({'success': True})

# Incoming stock
@app.route('/incoming')
@login_required
def incoming_stock():
    incoming = IncomingStock.query.filter_by(status='pending').all()
    products = Product.query.all()
    settings = Settings.query.first()
    return render_template('incoming.html', incoming=incoming, products=products, settings=settings)

@app.route('/incoming/add', methods=['POST'])
@login_required
def add_incoming():
    incoming = IncomingStock(
        product_id=request.form['product_id'],
        supplier=request.form['supplier'],
        expected_quantity=int(request.form['expected_quantity']),
        expected_date=datetime.strptime(request.form['expected_date'], '%Y-%m-%d').date(),
        po_number=request.form['po_number'],
        notes=request.form['notes'],
        created_by=current_user.id
    )
    db.session.add(incoming)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/incoming/<int:id>/receive', methods=['POST'])
@login_required
def receive_incoming(id):
    incoming = IncomingStock.query.get_or_404(id)
    if incoming.status != 'pending':
        return jsonify({'error': _('Invalid status')}), 400
    
    incoming.status = 'received'
    product = incoming.product
    product.quantity += incoming.expected_quantity
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/incoming/<int:id>/cancel', methods=['POST'])
@login_required
def cancel_incoming(id):
    incoming = IncomingStock.query.get_or_404(id)
    if incoming.status != 'pending':
        return jsonify({'error': _('Invalid status')}), 400
    
    incoming.status = 'cancelled'
    db.session.commit()
    return jsonify({'success': True})

@app.route('/system-admin')
@login_required
def system_admin():
    if current_user.role != 'admin':
        flash(_('Access denied. Admin privileges required.'), 'danger')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    return render_template('system_admin.html', users=users)

@app.route('/user/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        flash(_('Access denied. Admin privileges required.'), 'danger')
        return redirect(url_for('dashboard'))
    
    name = request.form.get('name')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')
    
    if User.query.filter_by(email=email).first():
        flash(_('Email already registered.'), 'danger')
        return redirect(url_for('system_admin'))
    
    user = User()
    user.name = name
    user.email = email
    user.role = role
    user.active = True
    user.set_password(password)
    
    db.session.add(user)
    db.session.commit()
    
    flash(_('User added successfully.'), 'success')
    return redirect(url_for('system_admin'))

@app.route('/user/<int:id>/edit', methods=['POST'])
@login_required
def edit_user(id):
    if current_user.role != 'admin':
        flash(_('Access denied. Admin privileges required.'), 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # Don't allow editing the last admin
    if user.role == 'admin' and user.id != current_user.id:
        admin_count = User.query.filter_by(role='admin').count()
        if admin_count <= 1 and request.form.get('role') != 'admin':
            flash(_('Cannot remove the last administrator.'), 'danger')
            return redirect(url_for('system_admin'))
    
    user.name = request.form.get('name')
    user.email = request.form.get('email')
    user.role = request.form.get('role')
    user.active = 'active' in request.form
    
    db.session.commit()
    
    flash(_('User updated successfully.'), 'success')
    return redirect(url_for('system_admin'))

@app.route('/user/<int:id>/delete', methods=['POST'])
@login_required
def delete_user(id):
    if current_user.role != 'admin':
        flash(_('Access denied. Admin privileges required.'), 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(id)
    
    # Don't allow deleting the last admin
    if user.role == 'admin':
        admin_count = User.query.filter_by(role='admin').count()
        if admin_count <= 1:
            flash(_('Cannot delete the last administrator.'), 'danger')
            return redirect(url_for('system_admin'))
    
    # Don't allow deleting yourself
    if user.id == current_user.id:
        flash(_('Cannot delete your own account.'), 'danger')
        return redirect(url_for('system_admin'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash(_('User deleted successfully.'), 'success')
    return redirect(url_for('system_admin'))

@app.route('/reset-admin', methods=['GET', 'POST'])
def reset_admin():
    if request.method == 'POST':
        # Check if admin exists
        admin = User.query.filter_by(email='admin@ewakor.com').first()
        if not admin:
            # Create admin if doesn't exist
            admin = User()
            admin.name = 'Admin'
            admin.email = 'admin@ewakor.com'
            admin.role = 'admin'
            admin.active = True
            admin.set_password('admin123')
            db.session.add(admin)
        
        # Set new password
        admin.set_password('admin123')
        db.session.commit()
        
        flash('Admin password has been reset to: admin123', 'success')
        return redirect(url_for('login'))
    
    return render_template('reset_admin.html')

def get_db_path():
    db_uri = app.config['SQLALCHEMY_DATABASE_URI']
    if db_uri.startswith('sqlite:///'):
        rel_path = db_uri.replace('sqlite:///', '', 1)
        return os.path.abspath(rel_path)
    return db_uri

@app.route('/backup')
@admin_required
def backup():
    db_path = get_db_path()
    backup_path = os.path.join(BACKUP_FOLDER, 'ewakor_backup.db')
    shutil.copy2(db_path, backup_path)
    return send_file(backup_path, as_attachment=True)

# Call this after adding/editing a product
def backup_db():
    db_path = get_db_path()
    backup_path = os.path.join(BACKUP_FOLDER, 'ewakor_autobackup.db')
    shutil.copy2(db_path, backup_path)

def save_image(file, folder):
    if file and file.filename:
        # Create a secure filename
        filename = secure_filename(file.filename)
        # Add timestamp to prevent filename collisions
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        
        # Save the file
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], folder, filename)
        file.save(file_path)
        
        # Process the image if it's a PNG
        if filename.lower().endswith('.png'):
            try:
                with Image.open(file_path) as img:
                    # Convert to RGBA if not already
                    if img.mode != 'RGBA':
                        img = img.convert('RGBA')
                    # Save the processed image
                    img.save(file_path, 'PNG')
            except Exception as e:
                print(f"Error processing image: {e}")
        
        # Return path relative to the static folder (e.g., 'uploads/logos/image.png')
        return os.path.join(app.config['UPLOAD_FOLDER'].replace('static/', ''), folder, filename).replace('\\', '/')
    return None

@app.route('/export/products')
@login_required
def export_products():
    export_format = request.args.get('format', 'csv')
    products = Product.query.all()
    headers = ['id', 'name', 'category', 'product_type', 'container_type', 'quantity', 'unit', 'image_path']
    if export_format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Produits'
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for p in products:
            ws.append([
                p.id,
                p.name,
                p.category,
                p.product_type,
                p.container_type,
                p.quantity,
                p.unit,
                p.image_path
            ])
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    else:
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(headers)
        for p in products:
            cw.writerow([
                p.id,
                p.name,
                p.category,
                p.product_type,
                p.container_type,
                p.quantity,
                p.unit,
                p.image_path
            ])
        output = si.getvalue().encode('utf-8')
        si.close()
        return send_file(
            BytesIO(output),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )

@app.route('/export/deliveries')
@login_required
def export_deliveries():
    export_format = request.args.get('format', 'csv')
    deliveries = IncomingStock.query.all()
    headers = ['id', 'product', 'supplier', 'expected_quantity', 'expected_date', 'status', 'po_number', 'notes']
    if export_format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Livraisons'
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for d in deliveries:
            ws.append([
                d.id,
                d.product.name if d.product else '',
                d.supplier,
                d.expected_quantity,
                d.expected_date,
                d.status,
                d.po_number,
                d.notes
            ])
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'deliveries_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    else:
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(headers)
        for d in deliveries:
            cw.writerow([
                d.id,
                d.product.name if d.product else '',
                d.supplier,
                d.expected_quantity,
                d.expected_date,
                d.status,
                d.po_number,
                d.notes
            ])
        output = si.getvalue().encode('utf-8')
        si.close()
        return send_file(
            BytesIO(output),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'deliveries_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )

@app.route('/backup/download')
@login_required
@admin_required
def download_backup():
    db_path = get_db_path()
    backup_name = f'ewakor_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
    backup_path = os.path.join(BACKUP_FOLDER, backup_name)
    shutil.copy2(db_path, backup_path)
    return send_file(backup_path, as_attachment=True)

@app.route('/backup/restore', methods=['POST'])
@login_required
@admin_required
def restore_backup():
    if 'backup_file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('dashboard'))
    file = request.files['backup_file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('dashboard'))
    filename = secure_filename(file.filename)
    restore_path = os.path.join(BACKUP_FOLDER, f'restore_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{filename}')
    file.save(restore_path)
    db_path = get_db_path()
    shutil.copy2(restore_path, db_path)
    flash('Backup restored successfully.', 'success')
    return redirect(url_for('dashboard'))

# Automatic daily backup logic

def daily_backup_thread():
    while True:
        now = datetime.now()
        next_run = now.replace(hour=0, minute=0, second=0, microsecond=0)
        if now >= next_run:
            next_run = next_run.replace(day=now.day + 1)
        wait_seconds = (next_run - now).total_seconds()
        time.sleep(wait_seconds)
        try:
            db_path = get_db_path()
            backup_name = f'ewakor_autobackup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
            backup_path = os.path.join(BACKUP_FOLDER, backup_name)
            shutil.copy2(db_path, backup_path)
        except Exception as e:
            print(f"Automatic backup failed: {e}")

# Start the backup thread only if running as main
if __name__ == '__main__':
    threading.Thread(target=daily_backup_thread, daemon=True).start()
    import sys
    host = '0.0.0.0'
    port = 5000
    if '--host=0.0.0.0' in sys.argv:
        host = '0.0.0.0'
    app.run(host=host, port=port, debug=True) 